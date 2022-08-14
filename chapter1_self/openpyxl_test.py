import openpyxl
from openpyxl.styles import Font


def create_xlsx():
    mywb = openpyxl.Workbook()
    sheet = mywb.active
    sheet.title = 'mynewtitle'
    print(mywb.sheetnames)
    mywb.save('mynewtitle.xlsx')


def copy_xlsx():
    mywb = openpyxl.load_workbook('mynewtitle.xlsx')
    sheet = mywb.active
    sheet.title = '副本'
    mywb.save('new.xlsx')


def create_sheet():
    mywb = openpyxl.Workbook()
    mywb.create_sheet()
    print(mywb.sheetnames)

    mywb.create_sheet(index=0, title='第一个工作簿')
    print(mywb.sheetnames)

    mywb.create_sheet(index=2, title='第二个工作簿')
    print(mywb.sheetnames)

    mywb.save('new.xlsx')


def del_sheet():
    mywb = openpyxl.load_workbook('new.xlsx')
    mywb.remove(mywb['Sheet1'])
    mywb.save('new.xlsx')


def insert_cell():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['Sheet']
    mysheet['F6'] = 'new value'
    mywb.save('new.xlsx')


def set_font():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['Sheet']
    font = Font(name='微软雅黑', size=32, italic=True)
    mysheet['F8'].font = font
    mysheet['F8'] = '你好，STYLES，蔡'
    mywb.save('new.xlsx')


def sum_cell():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['Sheet']
    mysheet['A1'] = 500
    mysheet['A2'] = 800
    mysheet['B1'] = '=SUM(A1:A2)'
    mywb.save('new.xlsx')

def set_row_height():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['Sheet']
    mysheet['C1'] = 'Tall row'
    mysheet['D2'] = 'Wide column'

    mysheet.row_dimensions[1].height = 65
    mysheet.column_dimensions['D'].width = 25
    mywb.save('new.xlsx')

def merge_cell():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['Sheet']
    mysheet.merge_cells('E2:G3')
    mysheet.merge_cells('I4:I8')
    mywb.save('new.xlsx')

def add_chart():
    mywb = openpyxl.load_workbook('new.xlsx')
    mysheet = mywb['第一个工作簿']

    for x in range(1, 10):
        mysheet['A' + str(x)] = x

    # 创建一个引用对象
    referenceobj = openpyxl.chart.Reference(mysheet, min_col=1, min_row=1, max_col=1, max_row=10)
    # 创建series对象，引用对象作为参数
    serObj = openpyxl.chart.Series(referenceobj, title='Chart Series')
    # 创建柱状图
    chartObject = openpyxl.chart.BarChart()
    # 将series对象作为参数
    chartObject.append(serObj)
    # 设置图表的宽和高
    chartObject.width = 20
    chartObject.height = 10
    # 在D2位置绘制图表
    mysheet.add_chart(chartObject, 'D2')
    mywb.save('new.xlsx')

def main():
    # copyXlsx()
    # createSheet()
    # del_sheet()
    # insert_cell()
    # set_font()
    # sum_cell()
    # set_row_height()
    # merge_cell()
    add_chart()


if __name__ == '__main__':
    main()
